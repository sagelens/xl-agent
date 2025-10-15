# excel_parser.py

import pandas as pd
import openpyxl
import re
from typing import Tuple
import os, sys 
sys.path.append(os.getcwd())
from src.database import DataDictionary

def clean_column_name(name: str) -> tuple[str, str | None]:
    if name is None: return "unnamed_col", None
    name_str = str(name).strip()
    
    unit = None
    unit_match = re.search(r'(.+?)\s*\((.+)\)$', name_str)
    if unit_match:
        name_str = unit_match.group(1).strip()
        unit = unit_match.group(2).strip()

    cleaned_name = name_str.lower()
    cleaned_name = re.sub(r'[^a-z0-9_]+', '_', cleaned_name)
    cleaned_name = cleaned_name.strip('_')
    
    return cleaned_name if cleaned_name else "unnamed_col", unit

def find_tables_on_sheet(sheet):
    tables = []
    visited_cells = set()

    for row_idx, row in enumerate(sheet.iter_rows()):
        for col_idx, cell in enumerate(row):
            cell_coord = (row_idx, col_idx)
            if cell.value is not None and cell_coord not in visited_cells:
                max_row, max_col = find_table_boundary(sheet, row_idx, col_idx)
                
                for r in range(row_idx, max_row + 1):
                    for c in range(col_idx, max_col + 1):
                        visited_cells.add((r, c))

                if max_row > row_idx or max_col > col_idx:
                    start_cell = openpyxl.utils.get_column_letter(col_idx + 1) + str(row_idx + 1)
                    end_cell = openpyxl.utils.get_column_letter(max_col + 1) + str(max_row + 1)
                    tables.append({"range": f"{start_cell}:{end_cell}"})
    return tables

def find_table_boundary(sheet, start_row, start_col):
    q = [(start_row, start_col)]
    visited = set(q)
    max_r, max_c = start_row, start_col

    while q:
        r, c = q.pop(0)
        max_r, max_c = max(max_r, r), max(max_c, c)
        
        for dr in range(-1, 2):
            for dc in range(-1, 2):
                if dr == 0 and dc == 0: continue
                
                nr, nc = r + dr, c + dc
                if 0 <= nr < sheet.max_row and 0 <= nc < sheet.max_column:
                    if (nr, nc) not in visited and sheet.cell(row=nr + 1, column=nc + 1).value is not None:
                        visited.add((nr, nc))
                        q.append((nr, nc))
    return max_r, max_c

def infer_and_clean_types(df, column_metadata=None):
    date_keywords = {'date', 'time', 'timestamp', 'datetime'}
    
    for col in df.columns:
        if col == '_provenance': continue
            
        col_str = str(col).lower()
        series = df[col].dropna()
        if series.empty: continue

        unit = None
        if column_metadata:
            for meta in column_metadata:
                if meta['name'] == col and meta.get('unit'):
                    unit = meta['unit'].lower()
                    break
        
        if any(kw in col_str for kw in date_keywords):
            if unit and any(term in unit for term in ['days', 'hours', 'minutes', 'duration']):
                df[col] = pd.to_numeric(df[col], errors='coerce')
                continue
            elif unit and any(term in unit for term in ['days', 'hours', 'duration']):
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(df[col])
                continue
            else:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                continue

        if 'id' in col_str:
            if series.astype(str).str.contains('[a-zA-Z]').any():
                continue

        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(df[col])
            
    return df

def parse_table(sheet, table_range, sheet_name) -> tuple[pd.DataFrame | None, dict | None]:
    data = sheet[table_range]
    rows_data = [[cell.value for cell in row] for row in data]
    df = pd.DataFrame(rows_data)

    if df.empty: return None, None

    header_row_index = 0
    for i in range(min(10, len(df))):
        row_values = df.iloc[i].dropna()
        if len(row_values) > 1 and row_values.apply(lambda x: isinstance(x, str)).mean() > 0.5:
            header_row_index = i
            break
    
    header_values = df.iloc[header_row_index]
    
    column_metadata = []
    new_columns = []
    for h in header_values:
        col_name, unit = clean_column_name(h)
        new_columns.append(col_name)
        column_metadata.append({"original_name": h, "name": col_name, "unit": unit})

    df.columns = new_columns
    
    cols = pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique():
        cols[cols[cols == dup].index.values.tolist()] = [dup + '_' + str(i) if i != 0 else dup for i in range(sum(cols == dup))]
    df.columns = cols
    
    data_df = df.iloc[header_row_index + 1:].reset_index(drop=True)
    if data_df.empty: return None, None
        
    data_df.dropna(how='all', inplace=True)

    try:
        min_col, min_row, max_col, _ = openpyxl.utils.cell.range_boundaries(table_range)
        start_col_letter = openpyxl.utils.get_column_letter(min_col)
        end_col_letter = openpyxl.utils.get_column_letter(max_col)
        
        data_start_row = min_row + header_row_index + 1
        
        provenance_list = [
            f"{sheet_name}!{start_col_letter}{data_start_row + i}:{end_col_letter}{data_start_row + i}"
            for i in range(len(data_df))
        ]
        data_df['_provenance'] = provenance_list
    except Exception:
        data_df['_provenance'] = f"Unknown range in {sheet_name}"
    
    data_df = infer_and_clean_types(data_df, column_metadata)
    return data_df, column_metadata

def process_workbook(file_path: str) -> tuple[dict[str, pd.DataFrame], DataDictionary]:
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return {}, DataDictionary()

    all_tables = {}
    data_dictionary = DataDictionary()

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            try:
                sheet.unmerge_cells(str(merged_range))
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_value = sheet.cell(row=min_row, column=min_col).value
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        sheet.cell(row=row, column=col).value = top_left_value
            except KeyError:
                continue
        
        table_definitions = find_tables_on_sheet(sheet)
        
        for i, t_def in enumerate(table_definitions):
            table_df, col_meta = parse_table(sheet, t_def['range'], sheet_name)
            if table_df is not None and not table_df.empty:
                table_name = f"{clean_column_name(sheet_name)[0]}_{i}"
                all_tables[table_name] = table_df
                
                data_dictionary.add_table(table_name)
                for j, col_name in enumerate(table_df.columns):
                    if col_name == '_provenance': continue
                    data_dictionary.add_column(
                        table_name=table_name,
                        col_name=col_name,
                        col_type=str(table_df[col_name].dtype),
                        unit=col_meta[j]['unit'] if j < len(col_meta) else None,
                        samples=table_df[col_name].dropna().head(3).tolist()
                    )
            
    return all_tables, data_dictionary